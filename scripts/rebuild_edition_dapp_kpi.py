#!/usr/bin/env python3
"""
rebuild_edition_dapp_kpi.py
---------------------------
Rebuilds public/data/dapp_kpi.json (Edition Demand & Collection page KPIs)
from public/data/edition_dapp.XLSX.

Classification rule (per instruction, 2 buckets only):
  CLP = milestone references a construction event (floor/slab, excavation,
        structure, OC, possession, flooring, finishing, plaster, basement,
        top floor), OR is a construction-event/fixed-date hybrid ("whichever
        is later/earlier" - the date is just a backstop on a real construction
        milestone), OR is a booking-stage milestone (Booking Amount/On
        Allotment - first installment, same convention as prior builds), OR
        is a fixed time offset FROM a construction event (e.g. "24 months
        from Application of OC").
  TLP = milestone is a pure calendar date or day/month interval with NO
        construction-event anchor (e.g. "On or Before 22 July 2026",
        "Within 60 Days From Date of Allotment").

Usage:
    python3 scripts/rebuild_edition_dapp_kpi.py
"""
import openpyxl
import json
import os
import re
from datetime import datetime, timedelta
from collections import defaultdict

BASE = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
FNAME = 'edition_dapp.XLSX'
OUT = 'dapp_kpi.json'

CONSTRUCTION_KW = ['floor', 'slab', 'excavation', 'structure', 'occupation certificate',
                   'occupat', 'possession', 'flooring', 'finishing work', 'plaster',
                   'top floor', 'basement', 'plinth']

MONTHS = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
          'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}


def extract_date(text):
    """Pull an explicit calendar date out of milestone text, e.g.
    'On or Before 15th May 2024', 'On or Before 28-01-2026', or
    'On or before 14th-Sep-2025'. Returns a datetime or None."""
    t = (text or '').replace('\xa0', ' ')
    m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?[\s-]+([A-Za-z]+)'?[\s,-]*(\d{4}|\d{2})\b", t)
    if m:
        mon_txt = m.group(2)[:3].lower()
        if mon_txt in MONTHS:
            year = int(m.group(3))
            if year < 100:
                year += 2000
            try:
                return datetime(year, MONTHS[mon_txt], int(m.group(1)))
            except ValueError:
                pass
    m2 = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', t)
    if m2:
        try:
            return datetime(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
        except ValueError:
            pass
    return None


def num(v):
    try:
        return float(v) if v not in (None, '') else 0.0
    except Exception:
        return 0.0


def load_rows():
    wb = openpyxl.load_workbook(os.path.join(BASE, FNAME), data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2):
        rows.append({headers[i]: row[i].value for i in range(len(headers))})
    return rows


def classify(milestone):
    n = (milestone or '').lower()
    has_constr = any(k in n for k in CONSTRUCTION_KW) or bool(re.search(r'\boc\b', n))
    if 'whichever is later' in n or 'whichever is earlier' in n:
        return 'clp'   # hybrid: construction event with a fixed-date backstop
    if has_constr:
        return 'clp'
    if 'booking amount' in n or n.strip() in ('on booking', 'on allotment'):
        return 'clp'   # booking-stage first installment
    if re.search(r'from application of', n) or re.search(r'from\s+occupat', n):
        return 'clp'   # time offset anchored to a construction event
    return 'tlp'        # pure fixed date or day/month interval from allotment/booking


def short_name(milestone, mtype):
    n = (milestone or '').strip()
    nl = n.lower()

    # floor/slab number
    m = re.search(r'(\d+)\s*(st|nd|rd|th)?\s*floor', nl)
    if m and 'flooring' not in nl:
        return f"{m.group(1)}{_ord_suffix(int(m.group(1)))} Floor Slab" if 'slab' in nl or True else f"{m.group(1)} Floor"
    if 'top floor' in nl:
        return 'Top Floor Slab'
    if 'occupation certificate' in nl or re.search(r'\boc\b', nl) or 'occupat' in nl:
        return 'OC Application'
    if 'possession' in nl:
        return 'Possession'
    if 'finishing work' in nl:
        return 'Finishing Work'
    if 'plaster' in nl:
        return 'External Plaster'
    if 'structure' in nl:
        return 'Structure Complete'
    if 'flooring' in nl:
        return 'Flooring'
    if 'basement' in nl:
        return 'Upper Basement'
    if 'plinth' in nl:
        return 'Plinth'
    if 'raft' in nl:
        return 'Raft Foundation'
    if 'excavation' in nl:
        return 'Excavation Complete' if 'complet' in nl else 'Excavation Start'
    if 'booking amount' in nl or nl in ('on booking',):
        return 'Booking Amount'
    if nl == 'on allotment':
        return 'On Allotment'
    # TLP "after" - interval from allotment/booking - keep each interval as its
    # own row (60 days is not the same milestone as 90 days)
    m2 = re.search(r'within\s+(\d+)\s*(days|months)', nl)
    if m2:
        anchor = 'Booking' if 'booking' in nl else 'Allotment'
        return f"After {m2.group(1)} {m2.group(2).title()} of {anchor}"
    # TLP "before" - a fixed calendar date, with or without an "execution/
    # signing" description - keep each date as its own row rather than
    # collapsing every fixed-date milestone into one generic bucket
    d = extract_date(n)
    if d:
        return f"Before {d.day} {d.strftime('%b %Y')}"
    # fallback: no recognizable date or keyword at all
    return n if len(n) <= 40 else n[:40].rstrip()


def _ord_suffix(n):
    if 11 <= n % 100 <= 13:
        return 'th'
    return {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')


def add_months(d, n):
    """Add n months to date d, clamping the day to the target month's length."""
    import calendar
    month = d.month - 1 + n
    year = d.year + month // 12
    month = month % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return datetime(year, month, day)


def construction_shortname(text):
    """Match ONLY the construction-event half of a milestone - no date/interval
    fallback - used to identify the 'construction side' of a hybrid milestone.
    Mirrors short_name()'s construction branches exactly."""
    nl = (text or '').lower()
    m = re.search(r'(\d+)\s*(st|nd|rd|th)?\s*floor', nl)
    if m and 'flooring' not in nl:
        return f"{m.group(1)}{_ord_suffix(int(m.group(1)))} Floor Slab"
    if 'top floor' in nl:
        return 'Top Floor Slab'
    if 'occupation certificate' in nl or re.search(r'\boc\b', nl) or 'occupat' in nl:
        return 'OC Application'
    if 'possession' in nl:
        return 'Possession'
    if 'finishing work' in nl or 'completion work' in nl:
        return 'Finishing Work'
    if 'plaster' in nl:
        return 'External Plaster'
    if 'structure' in nl:
        return 'Structure Complete'
    if 'flooring' in nl:
        return 'Flooring'
    if 'basement' in nl:
        return 'Upper Basement'
    if 'plinth' in nl:
        return 'Plinth'
    if 'raft' in nl:
        return 'Raft Foundation'
    if 'excavation' in nl:
        return 'Excavation Complete' if 'complet' in nl else 'Excavation Start'
    return None


def parse_relative_interval(text):
    """'within 36 months from allotment', '1 year from date of booking', etc.
    Returns (count, unit, anchor) or None."""
    nl = (text or '').lower()
    m = re.search(r'(\d+)\s*(day|month|year)s?\s+(?:from|of)\s+(?:date of\s+)?(allotment|booking)', nl)
    if m:
        return int(m.group(1)), m.group(2), m.group(3)
    return None


def parse_cell_date(v):
    """Parse a booking/allotment date cell (datetime object or string) to a plain datetime."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return datetime(v.year, v.month, v.day)
    try:
        return datetime.strptime(str(v)[:10], '%Y-%m-%d')
    except Exception:
        return None


def resolve_hybrid(milestone, row_booking_date, row_allotment_date, schedule_dates):
    """For a 'whichever is earlier/later' milestone, resolve THIS ROW'S actual
    winning date + classification by comparing:
      - the time-based side: that unit's OWN booking/allotment date + the
        stated interval (or a fixed date, when the text gives one instead of
        a relative interval)
      - the construction-event side: the project-wide schedule-file date for
        that construction event (same for every unit - schedule dates are not
        tracked per-tower)
    Returns (date, mtype, shortname) or None if the text isn't a hybrid
    pattern at all (caller should fall back to classify()/short_name()).
    """
    nl = (milestone or '').lower().replace('which ever', 'whichever')
    if 'whichever' not in nl:
        return None
    is_later = 'later' in nl

    time_date = None
    time_label = None
    rel = parse_relative_interval(milestone)
    if rel:
        count, unit, anchor = rel
        base = row_booking_date if anchor == 'booking' else row_allotment_date
        if base:
            if unit == 'day':
                time_date = base + timedelta(days=count)
            else:
                months = count if unit == 'month' else count * 12
                time_date = add_months(base, months)
        time_label = f"After {count} {unit.title()}{'s' if count != 1 else ''} of {anchor.title()}"
    else:
        fd = extract_date(milestone)
        if fd:
            time_date = fd
            time_label = f"Before {fd.day} {fd.strftime('%b %Y')}"

    constr_name = construction_shortname(milestone)
    constr_date = None
    if constr_name and schedule_dates.get(constr_name):
        try:
            constr_date = datetime.strptime(schedule_dates[constr_name], '%Y-%m')
        except Exception:
            constr_date = None

    if time_date is None and constr_date is None:
        return None
    if time_date is None:
        return (constr_date, 'clp', constr_name)
    if constr_date is None:
        return (time_date, 'tlp', time_label)

    winner = max(time_date, constr_date) if is_later else min(time_date, constr_date)
    if winner == time_date:
        return (time_date, 'tlp', time_label)
    return (constr_date, 'clp', constr_name)


# Slab-matrix-derived expected dates (MAX across towers = latest tower to
# reach that milestone), used only to fill milestones with NO Due Installment
# Date anywhere in the source DAPP file (per Edition_Slab_Matrix, AOP 30-06-2026).
SLAB_MATRIX_DATES = {
    'Excavation Start': '2024-02',
    'Excavation Complete': '2025-07',   # Ground Floor Slab completion per slab matrix
    '5th Floor Slab': '2025-12',
    '15th Floor Slab': '2026-05',
    '25th Floor Slab': '2026-09',
    '40th Floor Slab': '2027-04',
    'Structure Complete': '2027-03',
    'OC Application': '2027-12',
    # Not explicitly dated in the slab matrix (only 5th/15th/25th/top floor
    # + structure/OC are listed) - left blank rather than interpolated/guessed:
    #   20th Floor Slab, 29th Floor Slab, 32nd Floor Slab, 34th Floor Slab,
    #   Flooring, Finishing Work, Top Floor Slab, Possession ("After OC", no date)
}


def estimate_row_date(milestone, resolved_date, row_booking_date, row_allotment_date, sname, schedule_dates):
    """Best-effort date for an UNBILLED row, used to order each unit's future
    milestones chronologically for advance-netting and month bucketing.
    Priority: hybrid-resolved date > non-hybrid relative interval > non-hybrid
    fixed date > construction schedule-file date > None (unknown - can't be
    ordered, gets netted last)."""
    if resolved_date is not None:
        return resolved_date
    rel = parse_relative_interval(milestone)
    if rel:
        count, unit, anchor = rel
        base = row_booking_date if anchor == 'booking' else row_allotment_date
        if base:
            return base + timedelta(days=count) if unit == 'day' else add_months(base, count if unit == 'month' else count * 12)
    fd = extract_date(milestone)
    if fd:
        return fd
    if schedule_dates.get(sname):
        try:
            return datetime.strptime(schedule_dates[sname], '%Y-%m')
        except Exception:
            pass
    return None


def to_month(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return f"{v.year}-{v.month:02d}"
    try:
        s = str(v)[:10]
        d = datetime.strptime(s, '%Y-%m-%d')
        return f"{d.year}-{d.month:02d}"
    except Exception:
        return ''


def mlabel(m):
    try:
        return datetime.strptime(m, '%Y-%m').strftime("%b'%y")
    except Exception:
        return m


def main():
    print("Loading edition_dapp.XLSX...")
    rows = load_rows()
    print(f"  rows: {len(rows)}")

    buckets = {t: defaultdict(float) for t in ('all', 'tlp', 'clp')}
    tower_acc = defaultdict(lambda: defaultdict(float))
    monthly_acc = defaultdict(lambda: defaultdict(float))
    milestone_acc = defaultdict(lambda: defaultdict(float))
    milestone_type = {}
    milestone_tower = defaultdict(lambda: defaultdict(float))
    milestone_dates = defaultdict(list)

    adv_raw = {t: 0.0 for t in ('tlp', 'clp')}
    adv_raw_all = 0.0

    # Per-unit tracking for unbilled (demand==0) rows, so each unit's own
    # advance money can be netted off against ITS OWN future milestones in
    # chronological order (FIFO), rather than showing every future milestone
    # at its full undiscounted value regardless of money already on account.
    unit_advance_pool = defaultdict(float)
    unit_future_rows = defaultdict(list)  # unit -> [(date_or_None, installment, mtype, sname, tower)]

    today = datetime.today()
    current_month_start = datetime(today.year, today.month, 1)

    for r in rows:
        milestone = r.get('Milestone') or ''
        row_booking_date = parse_cell_date(r.get('SFDC Booking date'))
        row_allotment_date = parse_cell_date(r.get('Allotment Date'))
        hybrid = resolve_hybrid(milestone, row_booking_date, row_allotment_date, SLAB_MATRIX_DATES)
        resolved_date = None
        if hybrid:
            resolved_date, mtype, sname = hybrid
        else:
            mtype = classify(milestone)
            sname = short_name(milestone, mtype)

        demand = num(r.get('Demand Amount W/O Tax'))
        installment = num(r.get('Installment Amount'))
        bank = num(r.get('Received Amt (in Bank)'))
        cgst = num(r.get('CGST'))
        sgst = num(r.get('SGST'))
        recv_wot = bank - cgst - sgst
        outstanding = num(r.get('Outstanding 1'))
        tower = (r.get('Tower') or 'Unknown').strip() or 'Unknown'
        month = to_month(r.get('Bill creation date') or r.get('SAP Booking date'))
        due_month = to_month(r.get('Due Installment Date')) or (f"{resolved_date.year}-{resolved_date.month:02d}" if resolved_date else '')
        unit = (r.get('Unit Number') or '').strip()

        for t in ('all', mtype):
            buckets[t]['totalInstallment'] += demand
            buckets[t]['totalReceivedBank'] += bank
            buckets[t]['totalReceivedWoT'] += recv_wot
            buckets[t]['totalOutstanding'] += outstanding
            buckets[t]['totalDemandWoTax'] += demand

        # advance: billed demand not yet raised but money already received in bank
        if demand == 0 and bank > 0:
            adv_raw_all += bank
            adv_raw[mtype] += bank
            unit_advance_pool[unit] += bank

        # tower rollup (dem/rec/out split by type)
        tk = tower_acc[tower]
        tk[f'{mtype}_dem'] += demand
        tk[f'{mtype}_rec'] += recv_wot
        tk[f'{mtype}_out'] += outstanding

        # monthly rollup
        if month:
            mk = monthly_acc[month]
            mk[f'{mtype}_dem'] += demand
            mk[f'{mtype}_rec'] += recv_wot
            mk['outstanding'] += outstanding

        if demand > 0:
            # Already billed - this money is accounted for in the headline KPIs
            # already (Demand/Received/Outstanding above). It does NOT belong in
            # the forward-looking 'expected collection' milestone table/chart -
            # any still-unpaid portion is already captured by the single
            # Overdue/Outstanding entry added after this loop.
            pass
        else:
            # Not yet billed - hold for per-unit advance netting below.
            row_date = estimate_row_date(milestone, resolved_date, row_booking_date, row_allotment_date, sname, SLAB_MATRIX_DATES)
            unit_future_rows[unit].append([row_date, installment, mtype, sname, tower])

    # ---- Per-unit advance netting for unbilled (demand==0) rows ----
    # Each unit's own advance money (already collected, no bill raised yet) is
    # applied FIFO against ITS OWN future milestones in date order - a unit's
    # advance can only ever offset that same unit's own next bill, never
    # another unit's. Rows with no resolvable date are netted last (put at
    # the end of the queue) since they can't be chronologically ordered.
    # Past-due dates (the schedule/interval says it should already have
    # happened) are clamped up to the current month - that money is still
    # expected, just not still sitting in a stale past month.
    for unit, future_rows in unit_future_rows.items():
        future_rows.sort(key=lambda x: (x[0] is None, x[0] if x[0] else datetime.max))
        remaining_advance = unit_advance_pool.get(unit, 0.0)
        for row in future_rows:
            row_date, installment, mtype, sname, tower = row
            net_installment = installment
            if remaining_advance > 0:
                applied = min(remaining_advance, net_installment)
                net_installment -= applied
                remaining_advance -= applied
            clamped_date = row_date
            if clamped_date is not None and clamped_date < current_month_start:
                clamped_date = current_month_start
            due_month_net = f"{clamped_date.year}-{clamped_date.month:02d}" if clamped_date else ''

            milestone_acc[sname]['totalCr'] += net_installment
            milestone_acc[sname][f'{tower}'] = milestone_acc[sname].get(tower, 0) + net_installment
            milestone_type[sname] = mtype
            if due_month_net:
                milestone_dates[sname].append(due_month_net)

    # ---- Overdue/Outstanding: already-billed money still unpaid, folded into
    # the current month since it's expected to be collected now, not at
    # whatever original milestone date it was originally due -----------------
    if buckets['all']['totalOutstanding'] > 0:
        overdue_sname = 'Overdue / Outstanding Payments'
        current_month_key = f"{current_month_start.year}-{current_month_start.month:02d}"
        milestone_acc[overdue_sname]['totalCr'] += buckets['all']['totalOutstanding']
        milestone_type[overdue_sname] = 'overdue'
        milestone_dates[overdue_sname].append(current_month_key)


    def round_bucket(b):
        return {k: round(v / 1e7, 2) for k, v in b.items()}

    kpi = {t: round_bucket(buckets[t]) for t in buckets}

    GST_RATE = 0.05
    def adv_split(raw):
        net = raw / (1 + GST_RATE)
        gst = raw - net
        return round(raw / 1e7, 2), round(net / 1e7, 2), round(gst / 1e7, 2)

    adv_all_raw, adv_all_net, adv_all_gst = adv_split(adv_raw_all)
    advance_all = {'rawCr': adv_all_raw, 'netCr': adv_all_net, 'gstCr': adv_all_gst}
    advance = {}
    for t in ('tlp', 'clp'):
        rc, nc, gc = adv_split(adv_raw[t])
        advance[t] = {'rawCr': rc, 'netCr': nc, 'gstCr': gc}

    advance_note = (f"\u26a0\ufe0f Advance Money: \u20b9{adv_all_raw} Cr received before SAP demand was raised. "
                    f"GST @5% back-calculated (\u20b9{adv_all_gst} Cr) to show net W/O Tax of \u20b9{adv_all_net} Cr.")

    # ---- towerKpi ----
    tower_order = ['T-1', 'T-2', 'T-3', 'T-4', 'T-5', 'T-6']
    towerKpi = []
    for t in tower_order:
        v = tower_acc.get(t, {})
        towerKpi.append({
            'tower': t,
            'tlp_dem': round(v.get('tlp_dem', 0) / 1e7, 2),
            'clp_dem': round(v.get('clp_dem', 0) / 1e7, 2),
            'tlp_rec': round(v.get('tlp_rec', 0) / 1e7, 2),
            'clp_rec': round(v.get('clp_rec', 0) / 1e7, 2),
            'tlp_out': round(v.get('tlp_out', 0) / 1e7, 2),
            'clp_out': round(v.get('clp_out', 0) / 1e7, 2),
        })
    # any other towers not in the standard 6
    for t, v in tower_acc.items():
        if t not in tower_order:
            towerKpi.append({
                'tower': t,
                'tlp_dem': round(v.get('tlp_dem', 0) / 1e7, 2),
                'clp_dem': round(v.get('clp_dem', 0) / 1e7, 2),
                'tlp_rec': round(v.get('tlp_rec', 0) / 1e7, 2),
                'clp_rec': round(v.get('clp_rec', 0) / 1e7, 2),
                'tlp_out': round(v.get('tlp_out', 0) / 1e7, 2),
                'clp_out': round(v.get('clp_out', 0) / 1e7, 2),
            })

    # ---- monthlyTrend ----
    monthlyTrend = []
    for m in sorted(monthly_acc.keys()):
        v = monthly_acc[m]
        dem = v.get('tlp_dem', 0) + v.get('clp_dem', 0)
        rec = v.get('tlp_rec', 0) + v.get('clp_rec', 0)
        monthlyTrend.append({
            'month': m,
            'label': mlabel(m),
            'tlp_dem': round(v.get('tlp_dem', 0) / 1e7, 2),
            'clp_dem': round(v.get('clp_dem', 0) / 1e7, 2),
            'tlp_rec': round(v.get('tlp_rec', 0) / 1e7, 2),
            'clp_rec': round(v.get('clp_rec', 0) / 1e7, 2),
            'dem': round(dem / 1e7, 2),
            'rec': round(rec / 1e7, 2),
            'outstanding': round(v.get('outstanding', 0) / 1e7, 2),
        })

    # ---- milestonesUpcoming ----
    milestonesUpcoming = []
    for sname, acc in sorted(milestone_acc.items(), key=lambda x: -x[1]['totalCr']):
        mtype = milestone_type[sname]
        entry = {
            'name': sname,
            'type': mtype,
            'totalCr': round(acc['totalCr'] / 1e7, 2),
        }
        for t in tower_order:
            entry[t.replace('-', '')] = round(acc.get(t, 0) / 1e7, 2)
        dates = milestone_dates.get(sname, [])
        entry['expectedDate'] = max(dates) if dates else SLAB_MATRIX_DATES.get(sname, '')
        if mtype == 'overdue':
            entry['shortName'] = sname
        elif sname.startswith('Before ') or sname.startswith('After '):
            entry['shortName'] = sname  # already self-descriptive, no need for a redundant TLP prefix
        else:
            prefix = 'CLP' if mtype == 'clp' else 'TLP'
            entry['shortName'] = f"{prefix} \u2014 {sname}"
        milestonesUpcoming.append(entry)

    out = {
        'kpi': kpi,
        'gstRate': GST_RATE,
        'advanceNote': advance_note,
        'advance': advance,
        'advance_all': advance_all,
        'milestonesUpcoming': milestonesUpcoming,
        'monthlyTrend': monthlyTrend,
        'towerKpi': towerKpi,
    }

    # preserve projectMeta (unrelated to Edition; e.g. SKY ARC chip data lives here)
    existing_path = os.path.join(BASE, OUT)
    if os.path.exists(existing_path):
        existing = json.load(open(existing_path))
        out['projectMeta'] = existing.get('projectMeta', {})

    json.dump(out, open(existing_path, 'w'), separators=(',', ':'))

    print("\n\u2705 Done! dapp_kpi.json rebuilt.")
    print(f"  kpi.all: {kpi['all']}")
    print(f"  kpi.tlp: {kpi['tlp']}")
    print(f"  kpi.clp: {kpi['clp']}")
    print(f"  advance_all: {advance_all}")
    print(f"  milestone rows: {len(milestonesUpcoming)}")
    print(f"  monthly trend points: {len(monthlyTrend)}")


if __name__ == '__main__':
    main()
