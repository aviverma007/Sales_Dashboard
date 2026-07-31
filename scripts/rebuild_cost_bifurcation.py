#!/usr/bin/env python3
"""
rebuild_cost_bifurcation.py
---------------------------
Rebuilds public/data/cost_bifurcation_data.json from a ZALR export
(zalr_*.xlsx - one sheet, one header row, then rows grouped by
"WBS Description" - each group is several WBS detail rows followed by
one blank-WBS subtotal row for that group).

Source columns used:
  WBS, WBS Description, Plant Name, Budget, Total Actual, Assigned,
  Commitment, Budget Available, P.Grp Name (-> Department)

"WBS Description" (e.g. "House Keeping Expenses- FY-26-27") is the
Budget Head. The "- FY-26-27" suffix is stripped for display since every
row currently carries the same FY suffix.
"P.Grp Name" (Administration / IT) is the Department - added per instruction.

Usage:
    python3 scripts/rebuild_cost_bifurcation.py <path-to-zalr-file.xlsx>
"""
import openpyxl
import json
import os
import re
import sys
from collections import defaultdict

BASE = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
OUT = 'cost_bifurcation_data.json'


def num(v):
    try:
        return float(v) if v not in (None, '') else 0.0
    except Exception:
        return 0.0


def s(v):
    return str(v).strip() if v not in (None, '') else ''


def clean_budget_head(v):
    # Strip a trailing '- FY-26-27' / '-FY-26-27' style suffix (redundant -
    # every row in a single ZALR export carries the same FY).
    return re.sub(r'\s*-?\s*FY[-\s]?\d{2}[-\s]?\d{2}\s*$', '', s(v), flags=re.I).strip(' -')


def load_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows, idx


def agg_bucket(rows, key_field, label_field):
    """Group rows by key_field, summing budget/assigned/actual/commitment/available."""
    acc = defaultdict(lambda: defaultdict(float))
    counts = defaultdict(int)
    for r in rows:
        k = r[key_field]
        acc[k]['budget'] += r['budget']
        acc[k]['assigned'] += r['assigned']
        acc[k]['actual'] += r['actual']
        acc[k]['commitment'] += r['commitment']
        acc[k]['available'] += r['available']
        counts[k] += 1
    out = []
    for k, v in acc.items():
        util_pct = round(v['actual'] / v['budget'] * 100, 1) if v['budget'] > 0 else 0.0
        avail_pct = round(v['available'] / v['budget'] * 100, 1) if v['budget'] > 0 else 0.0
        out.append({
            label_field: k,
            'budget': round(v['budget'], 2),
            'assigned': round(v['assigned'], 2),
            'actual': round(v['actual'], 2),
            'commitment': round(v['commitment'], 2),
            'available': round(v['available'], 2),
            'budgetL': round(v['budget'] / 1e5, 2),
            'assignedL': round(v['assigned'] / 1e5, 2),
            'actualL': round(v['actual'] / 1e5, 2),
            'commitmentL': round(v['commitment'] / 1e5, 2),
            'availableL': round(v['available'] / 1e5, 2),
            'wbsCount': counts[k],
            'utilPct': util_pct,
            'availPct': avail_pct,
        })
    out.sort(key=lambda x: -x['budget'])
    return out


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else None
    if not src:
        print("Usage: python3 rebuild_cost_bifurcation.py <path-to-zalr-file.xlsx>")
        sys.exit(1)

    print(f"Loading {src} ...")
    raw_rows, idx = load_rows(src)

    # Detail rows = have a WBS code. Subtotal rows (blank WBS) are skipped -
    # we re-sum the detail rows ourselves rather than trust the source's own
    # subtotal rows, so the numbers are guaranteed to reconcile.
    detail = []
    for r in raw_rows:
        wbs = s(r.get('WBS'))
        if not wbs:
            continue
        budget_head_raw = s(r.get('WBS Description'))
        detail.append({
            'wbs': wbs,
            'budgetHeadRaw': budget_head_raw,
            'budgetHead': clean_budget_head(budget_head_raw),
            'site': s(r.get('Plant Name')),
            'department': s(r.get('P.Grp Name')) or 'Unspecified',
            'description': budget_head_raw,
            'budget': num(r.get('Budget')),
            'assigned': num(r.get('Assigned')),
            'actual': num(r.get('Total Actual')),
            'commitment': num(r.get('Commitment')),
            'available': num(r.get('Budget Available')),
        })

    print(f"  detail rows (with WBS code): {len(detail)}")

    total_budget = sum(r['budget'] for r in detail)
    total_assigned = sum(r['assigned'] for r in detail)
    total_actual = sum(r['actual'] for r in detail)
    total_commitment = sum(r['commitment'] for r in detail)
    total_available = sum(r['available'] for r in detail)

    kpi = {
        'totalBudgetL': round(total_budget / 1e5, 2),
        'totalAssignedL': round(total_assigned / 1e5, 2),
        'totalActualL': round(total_actual / 1e5, 2),
        'totalCommitmentL': round(total_commitment / 1e5, 2),
        'totalAvailableL': round(total_available / 1e5, 2),
        'utilizationPct': round(total_actual / total_budget * 100, 1) if total_budget > 0 else 0,
        'wbsCount': len(detail),
    }

    by_budget_head = agg_bucket(detail, 'budgetHead', 'budgetHead')
    by_department = agg_bucket(detail, 'department', 'department')

    wbs_table = []
    for r in sorted(detail, key=lambda x: -x['budget']):
        util_pct = round(r['actual'] / r['budget'] * 100, 1) if r['budget'] > 0 else 0.0
        wbs_table.append({
            'site': r['site'],
            'wbs': r['wbs'],
            'budgetHead': r['budgetHead'],
            'department': r['department'],
            'description': r['description'],
            'budget': round(r['budget'], 2),
            'assigned': round(r['assigned'], 2),
            'actual': round(r['actual'], 2),
            'commitment': round(r['commitment'], 2),
            'available': round(r['available'], 2),
            'budgetL': round(r['budget'] / 1e5, 2),
            'assignedL': round(r['assigned'] / 1e5, 2),
            'actualL': round(r['actual'] / 1e5, 2),
            'commitmentL': round(r['commitment'] / 1e5, 2),
            'availableL': round(r['available'] / 1e5, 2),
            'utilPct': util_pct,
        })

    filter_options = {
        'budgetHeads': sorted(set(r['budgetHead'] for r in detail)),
        'departments': sorted(set(r['department'] for r in detail)),
    }

    out = {
        'kpi': kpi,
        'byBudgetHead': by_budget_head,
        'byDepartment': by_department,
        'wbsTable': wbs_table,
        'filterOptions': filter_options,
        'meta': {
            'source': os.path.basename(src),
            'rebuiltBy': 'scripts/rebuild_cost_bifurcation.py',
        },
    }

    out_path = os.path.join(BASE, OUT)
    json.dump(out, open(out_path, 'w'), separators=(',', ':'))

    print("\n\u2705 Done! cost_bifurcation_data.json rebuilt.")
    print(f"  Total Budget: \u20b9{kpi['totalBudgetL']} L | Actual: \u20b9{kpi['totalActualL']} L | Utilization: {kpi['utilizationPct']}%")
    print(f"  Budget Heads: {len(by_budget_head)} | Departments: {len(by_department)} | WBS rows: {len(wbs_table)}")


if __name__ == '__main__':
    main()
