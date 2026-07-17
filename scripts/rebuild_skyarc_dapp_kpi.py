#!/usr/bin/env python3
"""
rebuild_skyarc_dapp_kpi.py
--------------------------
Rebuilds public/data/skyarc_dapp_kpi.json (Sky Arc Demand & Collection page KPIs)
from public/data/skyarc_dapp.XLSX.

Classification is now by MILESTONE CONTENT (same rule as Edition), not by
Payment Plan Name - this guarantees each milestone appears in exactly one
row, since a milestone's meaning doesn't change depending on which plan
happened to bill it. See rebuild_edition_dapp_kpi.py for the full rationale.

  CLP = construction event (floor/slab, excavation, structure, OC,
        possession, flooring, finishing, plaster, basement/plinth/top
        floor), OR hybrid event-or-date ("whichever is later/earlier" -
        date is just a backstop on a real construction milestone), OR
        booking-stage (Booking Amount/On Allotment), OR a time offset
        anchored to a construction event.
  TLP  = "Before <date>" - a pure fixed calendar date with no construction
         anchor. Each distinct date is its OWN row - never collapsed into
         one generic bucket.
       = "After X Days/Months of Allotment/Booking" - a relative interval
         with no construction anchor. Each distinct interval is its own row.

Tower mapping for the milestone table: TA->T1, TB->T2, TC->T3, TD->T4,
TE->T5, TF->T6 (fixed positional mapping, per doc).

Usage:
    python3 scripts/rebuild_skyarc_dapp_kpi.py
"""
import openpyxl
import json
import os
import re
from datetime import datetime
from collections import defaultdict

BASE = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
FNAME = 'skyarc_dapp.XLSX'
OUT = 'skyarc_dapp_kpi.json'

TOWER_MAP = {'TA': 'T1', 'TB': 'T2', 'TC': 'T3', 'TD': 'T4', 'TE': 'T5', 'TF': 'T6'}

CONSTRUCTION_KW = ['floor', 'slab', 'excavation', 'structure', 'occupation certificate',
                   'occupat', 'possession', 'flooring', 'finishing work', 'plaster',
                   'top floor', 'basement', 'plinth']

MONTHS = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
          'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}


def num(v):
    try:
        return float(v) if v not in (None, '') else 0.0
    except Exception:
        return 0.0


def load_rows():
    wb = openpyxl.load_workbook(os.path.join(BASE, FNAME), data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2):
        rows.append({headers[i]: row[i].value for i in range(len(headers))})
    return rows


def extract_date(text):
    t = (text or '').replace(' ', ' ')
    m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?[\s-]+([A-Za-z]+)'?[\s-]*(\d{4}|\d{2})\b", t)
    if m:
        mon_txt = m.group(2)[:3].lower()
        if mon_txt in MONTHS:
            year = int(m.group(3))
            if year < 100:
                year += 2000
            try:
                return datetime(year, MONTHS[mon_txt], int(m.group(1)))
            except ValueError:
                pass
    m2 = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', t)
    if m2:
        try:
            return datetime(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
        except ValueError:
            pass
    return None


def classify(milestone):
    n = (milestone or '').lower()
    has_constr = any(k in n for k in CONSTRUCTION_KW) or bool(re.search(r'\boc\b', n))
    if 'whichever is later' in n or 'whichever is earlier' in n:
        return 'clp'
    if has_constr:
        return 'clp'
    if 'booking amount' in n or n.strip() in ('on booking', 'on allotment'):
        return 'clp'
    if re.search(r'from application of', n) or re.search(r'from\s+occupat', n):
        return 'clp'
    return 'tlp'


def short_name(milestone):
    n = (milestone or '').strip()
    nl = n.lower()

    m = re.search(r'(\d+)\s*(st|nd|rd|th)?\s*floor', nl)
    if m and 'flooring' not in nl:
        return f"{m.group(1)}{_ord_suffix(int(m.group(1)))} Floor Slab"
    if 'top floor' in nl:
        return 'Top Floor Slab'
    if 'occupation certificate' in nl or re.search(r'\boc\b', nl) or 'occupat' in nl:
        return 'OC Application'
    if 'possession' in nl:
        return 'Possession'
    if 'finishing work' in nl:
        return 'Finishing Work'
    if 'plaster' in nl:
        return 'External Plaster'
    if 'structure' in nl:
        return 'Structure Complete'
    if 'flooring' in nl:
        return 'Flooring'
    if 'basement' in nl:
        return 'Upper Basement'
    if 'plinth' in nl:
        return 'Plinth'
    if 'excavation' in nl:
        return 'Excavation Complete' if 'complet' in nl else 'Excavation Start'
    if 'booking amount' in nl or nl in ('on booking',):
        return 'Booking Amount'
    if nl == 'on allotment':
        return 'On Allotment'
    m2 = re.search(r'within\s+(\d+)\s*(days|months)', nl)
    if m2:
        anchor = 'Booking' if 'booking' in nl else 'Allotment'
        return f"After {m2.group(1)} {m2.group(2).title()} of {anchor}"
    d = extract_date(n)
    if d:
        return f"Before {d.day} {d.strftime('%b %Y')}"
    return n if len(n) <= 40 else n[:40].rstrip()


def _ord_suffix(n):
    if 11 <= n % 100 <= 13:
        return 'th'
    return {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')


# Dates from Sky Arc schedule (11__Sky_Arc_schedule_-_Update-01_Jul_26_.xlsx,
# 'Summary' sheet, CRM Milestone Completions table). Using the latest tower's
# actual/CRM date, same MAX-across-towers convention as Edition's slab matrix.
SLAB_SCHEDULE_DATES = {
    '40th Floor Slab': '2027-10',
    '30th Floor Slab': '2027-04',
    '20th Floor Slab': '2026-11',
    '5th Floor Slab': '2026-07',
    'External Plaster': '2028-04',
    'Flooring': '2028-11',
}


def to_month(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return f"{v.year}-{v.month:02d}"
    try:
        s = str(v)[:10]
        d = datetime.strptime(s, '%Y-%m-%d')
        return f"{d.year}-{d.month:02d}"
    except Exception:
        return ''


def mlabel(m):
    try:
        return datetime.strptime(m, '%Y-%m').strftime("%b'%y")
    except Exception:
        return m


def main():
    print("Loading skyarc_dapp.XLSX...")
    rows = load_rows()
    print(f"  rows: {len(rows)}")

    buckets = {t: defaultdict(float) for t in ('all', 'tlp', 'clp')}
    tower_acc = defaultdict(lambda: defaultdict(float))
    monthly_acc = defaultdict(lambda: defaultdict(float))
    milestone_acc = defaultdict(lambda: defaultdict(float))
    milestone_type = {}
    milestone_dates = defaultdict(list)

    adv_raw = {t: 0.0 for t in ('tlp', 'clp')}
    adv_raw_all = 0.0

    for r in rows:
        milestone = (r.get('Milestone') or '').strip()
        mtype = classify(milestone)
        sname = short_name(milestone)

        demand = num(r.get('Demand Amount W/O Tax'))
        installment = num(r.get('Installment Amount'))
        bank = num(r.get('Received Amt (in Bank)'))
        cgst = num(r.get('CGST'))
        sgst = num(r.get('SGST'))
        recv_wot = bank - cgst - sgst
        outstanding = num(r.get('Outstanding 1'))
        tower_raw = (r.get('Tower') or 'Unknown').strip() or 'Unknown'
        month = to_month(r.get('Bill creation date') or r.get('SAP Booking date'))
        due_month = to_month(r.get('Due Installment Date'))

        for t in ('all', mtype):
            buckets[t]['totalInstallment'] += demand
            buckets[t]['totalReceivedBank'] += bank
            buckets[t]['totalReceivedWoT'] += recv_wot
            buckets[t]['totalOutstanding'] += outstanding
            buckets[t]['totalDemandWoTax'] += demand

        if demand == 0 and bank > 0:
            adv_raw_all += bank
            adv_raw[mtype] += bank

        tk = tower_acc[tower_raw]
        tk[f'{mtype}_dem'] += demand
        tk[f'{mtype}_rec'] += recv_wot
        tk[f'{mtype}_out'] += outstanding

        if month:
            mk = monthly_acc[month]
            mk[f'{mtype}_dem'] += demand
            mk[f'{mtype}_rec'] += recv_wot

        # milestone rollup - grouped by shortName (semantic, content-derived),
        # not raw text or plan type - guarantees one row per unique milestone
        acc = milestone_acc[sname]
        acc['totalCr'] += installment
        acc[tower_raw] = acc.get(tower_raw, 0) + installment
        milestone_type[sname] = mtype
        if due_month:
            milestone_dates[sname].append(due_month)

    def round_bucket(b):
        return {k: round(v / 1e7, 2) for k, v in b.items()}

    kpi = {t: round_bucket(buckets[t]) for t in buckets}

    GST_RATE = 0.05
    def adv_split(raw):
        net = raw / (1 + GST_RATE)
        gst = raw - net
        return round(raw / 1e7, 2), round(net / 1e7, 2), round(gst / 1e7, 2)

    adv_all_raw, adv_all_net, adv_all_gst = adv_split(adv_raw_all)
    advance_all = {'rawCr': adv_all_raw, 'netCr': adv_all_net, 'gstCr': adv_all_gst}
    advance = {}
    for t in ('tlp', 'clp'):
        rc, nc, gc = adv_split(adv_raw[t])
        advance[t] = {'rawCr': rc, 'netCr': nc, 'gstCr': gc}

    advance_note = (f"\u26a0\ufe0f Advance Money: \u20b9{adv_all_raw} Cr received before SAP demand was raised. "
                    f"GST @5% back-calculated (\u20b9{adv_all_gst} Cr) to show net W/O Tax of \u20b9{adv_all_net} Cr.")

    tower_order = ['TA', 'TB', 'TC', 'TD', 'TE', 'TF']
    towerKpi = []
    for t in tower_order:
        v = tower_acc.get(t, {})
        towerKpi.append({
            'tower': t,
            'tlp_dem': round(v.get('tlp_dem', 0) / 1e7, 2),
            'clp_dem': round(v.get('clp_dem', 0) / 1e7, 2),
            'tlp_rec': round(v.get('tlp_rec', 0) / 1e7, 2),
            'clp_rec': round(v.get('clp_rec', 0) / 1e7, 2),
            'tlp_out': round(v.get('tlp_out', 0) / 1e7, 2),
            'clp_out': round(v.get('clp_out', 0) / 1e7, 2),
        })
    for t, v in tower_acc.items():
        if t not in tower_order:
            towerKpi.append({
                'tower': t,
                'tlp_dem': round(v.get('tlp_dem', 0) / 1e7, 2),
                'clp_dem': round(v.get('clp_dem', 0) / 1e7, 2),
                'tlp_rec': round(v.get('tlp_rec', 0) / 1e7, 2),
                'clp_rec': round(v.get('clp_rec', 0) / 1e7, 2),
                'tlp_out': round(v.get('tlp_out', 0) / 1e7, 2),
                'clp_out': round(v.get('clp_out', 0) / 1e7, 2),
            })

    monthlyTrend = []
    for m in sorted(monthly_acc.keys()):
        v = monthly_acc[m]
        dem = v.get('tlp_dem', 0) + v.get('clp_dem', 0)
        rec = v.get('tlp_rec', 0) + v.get('clp_rec', 0)
        monthlyTrend.append({
            'month': m,
            'label': mlabel(m),
            'tlp_dem': round(v.get('tlp_dem', 0) / 1e7, 2),
            'clp_dem': round(v.get('clp_dem', 0) / 1e7, 2),
            'tlp_rec': round(v.get('tlp_rec', 0) / 1e7, 2),
            'clp_rec': round(v.get('clp_rec', 0) / 1e7, 2),
            'dem': round(dem / 1e7, 2),
            'rec': round(rec / 1e7, 2),
        })

    milestonesUpcoming = []
    for sname, acc in sorted(milestone_acc.items(), key=lambda x: -x[1]['totalCr']):
        mtype = milestone_type[sname]
        entry = {
            'name': sname,
            'type': mtype,
            'totalCr': round(acc['totalCr'] / 1e7, 2),
        }
        for raw_t, mapped_t in TOWER_MAP.items():
            entry[mapped_t] = round(acc.get(raw_t, 0) / 1e7, 2)
        dates = milestone_dates.get(sname, [])
        entry['expectedDate'] = max(dates) if dates else SLAB_SCHEDULE_DATES.get(sname, '')
        prefix = 'CLP' if mtype == 'clp' else 'TLP'
        if sname.startswith('Before ') or sname.startswith('After '):
            entry['shortName'] = sname
        else:
            entry['shortName'] = f"{prefix} \u2014 {sname}"
        milestonesUpcoming.append(entry)

    out = {
        'kpi': kpi,
        'gstRate': GST_RATE,
        'advanceNote': advance_note,
        'advance': advance,
        'advance_all': advance_all,
        'milestonesUpcoming': milestonesUpcoming,
        'monthlyTrend': monthlyTrend,
        'towerKpi': towerKpi,
    }

    existing_path = os.path.join(BASE, OUT)
    if os.path.exists(existing_path):
        existing = json.load(open(existing_path))
        if existing.get('projectMeta'):
            out['projectMeta'] = existing['projectMeta']

    json.dump(out, open(existing_path, 'w'), separators=(',', ':'))

    print("\n\u2705 Done! skyarc_dapp_kpi.json rebuilt.")
    print(f"  kpi.all: {kpi['all']}")
    print(f"  kpi.tlp: {kpi['tlp']}")
    print(f"  kpi.clp: {kpi['clp']}")
    print(f"  advance_all: {advance_all}")
    print(f"  milestone rows: {len(milestonesUpcoming)}")
    print(f"  monthly trend points: {len(monthlyTrend)}")


if __name__ == '__main__':
    main()
