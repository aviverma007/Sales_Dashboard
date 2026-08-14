#!/usr/bin/env python3
"""
rebuild_trump_dapp_kpi.py
--------------------------
Rebuilds public/data/trump_dapp_kpi.json (Trump Demand & Collection page KPIs)
from public/data/trump_dapp.XLSX.

Classification is now by MILESTONE CONTENT (same rule as Edition/Sky Arc),
superseding the old "all TLP" convention. Trump's Payment Plan Name values
are ratio-style ("30:30:40", "35:65") and never contain "CLP"/"TLP" text,
so Payment Plan Name was never a usable signal - but the milestone text
itself does describe real construction events (floor slabs, structure,
OC, possession), so content-based classification works the same as it
does for Edition/Sky Arc.

  CLP = construction event, hybrid event-or-date, booking-stage, or a time
        offset anchored to a construction event.
  TLP = "Before <date>" - a pure fixed calendar date, each distinct date
        its own row.
      = "After X Days/Months of Allotment/Booking" - a relative interval,
        each distinct interval its own row.

Towers: Tower-1, Tower-2, plus a combined "Tower-1 & 2" bucket (not split
into T-columns, per doc). The milestone table's T1-T6 columns are hardcoded
in the frontend, so Tower-1->T1, Tower-2->T2; the combined bucket
contributes to the milestone's totalCr but isn't attributed to a specific
T-column (T3-T6 stay 0 for Trump).

Usage:
    python3 scripts/rebuild_trump_dapp_kpi.py
"""
import openpyxl
import json
import os
import re
from datetime import datetime, timedelta
from collections import defaultdict

BASE = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
FNAME = 'trump_dapp.XLSX'
OUT = 'trump_dapp_kpi.json'

TOWER_MAP = {'Tower-1': 'T1', 'Tower-2': 'T2'}

CONSTRUCTION_KW = ['floor', 'slab', 'excavation', 'structure', 'occupation certificate',
                   'occupat', 'possession', 'flooring', 'finishing work', 'plaster',
                   'top floor', 'basement', 'plinth']

MONTHS = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
          'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}


def num(v):
    try:
        return float(v) if v not in (None, '') else 0.0
    except Exception:
        return 0.0


def load_rows():
    wb = openpyxl.load_workbook(os.path.join(BASE, FNAME), data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2):
        rows.append({headers[i]: row[i].value for i in range(len(headers))})
    return rows


def extract_date(text):
    t = (text or '').replace('\xa0', ' ')
    m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?[\s-]+([A-Za-z]+)'?[\s,-]*(\d{4}|\d{2})\b", t)
    if m:
        mon_txt = m.group(2)[:3].lower()
        if mon_txt in MONTHS:
            year = int(m.group(3))
            if year < 100:
                year += 2000
            try:
                return datetime(year, MONTHS[mon_txt], int(m.group(1)))
            except ValueError:
                pass
    m2 = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', t)
    if m2:
        try:
            return datetime(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
        except ValueError:
            pass
    return None


def classify(milestone):
    n = (milestone or '').lower()
    has_constr = any(k in n for k in CONSTRUCTION_KW) or bool(re.search(r'\boc\b', n))
    if 'whichever is later' in n or 'whichever is earlier' in n:
        return 'clp'
    if has_constr:
        return 'clp'
    if 'booking amount' in n or n.strip() in ('on booking', 'on allotment'):
        return 'clp'
    if re.search(r'from application of', n) or re.search(r'from\s+occupat', n):
        return 'clp'
    return 'tlp'


def short_name(milestone):
    n = (milestone or '').strip()
    nl = n.lower()

    m = re.search(r'(\d+)\s*(st|nd|rd|th)?\s*floor', nl)
    if m and 'flooring' not in nl:
        return f"{m.group(1)}{_ord_suffix(int(m.group(1)))} Floor Slab"
    if 'top floor' in nl:
        return 'Top Floor Slab'
    if 'occupation certificate' in nl or re.search(r'\boc\b', nl) or 'occupat' in nl:
        return 'OC Application'
    if 'possession' in nl:
        return 'Possession'
    if 'finishing work' in nl:
        return 'Finishing Work'
    if 'plaster' in nl:
        return 'External Plaster'
    if 'structure' in nl:
        return 'Structure Complete'
    if 'flooring' in nl:
        return 'Flooring'
    if 'basement' in nl:
        return 'Upper Basement'
    if 'plinth' in nl:
        return 'Plinth'
    if 'raft' in nl:
        return 'Raft Foundation'
    if 'excavation' in nl:
        return 'Excavation Complete' if 'complet' in nl else 'Excavation Start'
    if 'booking amount' in nl or nl in ('on booking',):
        return 'Booking Amount'
    if nl == 'on allotment':
        return 'On Allotment'
    m2 = re.search(r'within\s+(\d+)\s*(days|months)', nl)
    if m2:
        anchor = 'Booking' if 'booking' in nl else 'Allotment'
        return f"After {m2.group(1)} {m2.group(2).title()} of {anchor}"
    d = extract_date(n)
    if d:
        return f"Before {d.day} {d.strftime('%b %Y')}"
    return n if len(n) <= 40 else n[:40].rstrip()


def _ord_suffix(n):
    if 11 <= n % 100 <= 13:
        return 'th'
    return {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')


def add_months(d, n):
    """Add n months to date d, clamping the day to the target month's length."""
    import calendar
    month = d.month - 1 + n
    year = d.year + month // 12
    month = month % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return datetime(year, month, day)


def construction_shortname(text):
    """Match ONLY the construction-event half of a milestone - no date/interval
    fallback - used to identify the 'construction side' of a hybrid milestone."""
    nl = (text or '').lower()
    m = re.search(r'(\d+)\s*(st|nd|rd|th)?\s*floor', nl)
    if m and 'flooring' not in nl:
        return f"{m.group(1)}{_ord_suffix(int(m.group(1)))} Floor Slab"
    if 'top floor' in nl:
        return 'Top Floor Slab'
    if 'occupation certificate' in nl or re.search(r'\boc\b', nl) or 'occupat' in nl:
        return 'OC Application'
    if 'possession' in nl:
        return 'Possession'
    if 'finishing work' in nl or 'completion work' in nl:
        return 'Finishing Work'
    if 'plaster' in nl:
        return 'External Plaster'
    if 'structure' in nl:
        return 'Structure Complete'
    if 'flooring' in nl:
        return 'Flooring'
    if 'basement' in nl:
        return 'Upper Basement'
    if 'plinth' in nl:
        return 'Plinth'
    if 'raft' in nl:
        return 'Raft Foundation'
    if 'excavation' in nl:
        return 'Excavation Complete' if 'complet' in nl else 'Excavation Start'
    return None


def parse_relative_interval(text):
    """'36th months of booking', '1 year from date of booking', etc.
    Returns (count, unit, anchor) or None."""
    nl = (text or '').lower()
    m = re.search(r'(\d+)\s*(?:st|nd|rd|th)?\s*(day|month|year)s?\s+(?:from|of)\s+(?:date of\s+)?(allotment|booking)', nl)
    if m:
        return int(m.group(1)), m.group(2), m.group(3)
    return None


def parse_cell_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return datetime(v.year, v.month, v.day)
    try:
        return datetime.strptime(str(v)[:10], '%Y-%m-%d')
    except Exception:
        return None


def resolve_hybrid(milestone, row_booking_date, row_allotment_date, schedule_dates):
    """For a 'whichever is earlier/later' milestone, resolve THIS ROW'S actual
    winning date + classification: compare that unit's own booking/allotment
    date + stated interval (or a fixed date, if the text gives one instead of
    a relative interval) against the project-wide schedule-file date for the
    construction event. Returns (date, mtype, shortname) or None if the text
    isn't a hybrid pattern at all."""
    nl = (milestone or '').lower().replace('which ever', 'whichever')
    if 'whichever' not in nl:
        return None
    is_later = 'later' in nl

    time_date = None
    time_label = None
    rel = parse_relative_interval(milestone)
    if rel:
        count, unit, anchor = rel
        base = row_booking_date if anchor == 'booking' else row_allotment_date
        if base:
            if unit == 'day':
                time_date = base + timedelta(days=count)
            else:
                months = count if unit == 'month' else count * 12
                time_date = add_months(base, months)
        time_label = f"After {count} {unit.title()}{'s' if count != 1 else ''} of {anchor.title()}"
    else:
        fd = extract_date(milestone)
        if fd:
            time_date = fd
            time_label = f"Before {fd.day} {fd.strftime('%b %Y')}"

    constr_name = construction_shortname(milestone)
    constr_date = None
    if constr_name and schedule_dates.get(constr_name):
        try:
            constr_date = datetime.strptime(schedule_dates[constr_name], '%Y-%m')
        except Exception:
            constr_date = None

    if time_date is None and constr_date is None:
        return None
    if time_date is None:
        return (constr_date, 'clp', constr_name)
    if constr_date is None:
        return (time_date, 'tlp', time_label)

    winner = max(time_date, constr_date) if is_later else min(time_date, constr_date)
    if winner == time_date:
        return (time_date, 'tlp', time_label)
    return (constr_date, 'clp', constr_name)


# Dates from Trump Tower schedule (20260702_Trump_Tower_L0.xlsx, '31.10.2025'
# sheet), MAX across Tower-1/Tower-2 finish dates, same convention as
# Edition/Sky Arc.
SLAB_SCHEDULE_DATES = {
    '40th Floor Slab': '2028-06',
    '10th Floor Slab': '2027-03',
    '25th Floor Slab': '2027-11',
    'Structure Complete': '2029-01',  # LMR/Mumty Roof Slab used as proxy
    'Finishing Work': '2029-09',
    'OC Application': '2029-10',
    'Possession': '2029-10',  # Hand over date used as possession proxy
}


def estimate_row_date(milestone, resolved_date, row_booking_date, row_allotment_date, sname, schedule_dates):
    """Best-effort date for an UNBILLED row, used to order each unit's future
    milestones chronologically for advance-netting and month bucketing."""
    if resolved_date is not None:
        return resolved_date
    rel = parse_relative_interval(milestone)
    if rel:
        count, unit, anchor = rel
        base = row_booking_date if anchor == 'booking' else row_allotment_date
        if base:
            return base + timedelta(days=count) if unit == 'day' else add_months(base, count if unit == 'month' else count * 12)
    fd = extract_date(milestone)
    if fd:
        return fd
    if schedule_dates.get(sname):
        try:
            return datetime.strptime(schedule_dates[sname], '%Y-%m')
        except Exception:
            pass
    return None


def to_month(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return f"{v.year}-{v.month:02d}"
    try:
        s = str(v)[:10]
        d = datetime.strptime(s, '%Y-%m-%d')
        return f"{d.year}-{d.month:02d}"
    except Exception:
        return ''


def mlabel(m):
    try:
        return datetime.strptime(m, '%Y-%m').strftime("%b'%y")
    except Exception:
        return m


def main():
    print("Loading trump_dapp.XLSX...")
    rows = load_rows()
    print(f"  rows: {len(rows)}")

    buckets = {t: defaultdict(float) for t in ('all', 'tlp', 'clp')}
    tower_acc = defaultdict(lambda: defaultdict(float))
    monthly_acc = defaultdict(lambda: defaultdict(float))
    milestone_acc = defaultdict(lambda: defaultdict(float))
    milestone_type = {}
    milestone_dates = defaultdict(list)

    adv_raw = {t: 0.0 for t in ('tlp', 'clp')}
    adv_raw_all = 0.0

    unit_advance_pool = defaultdict(float)
    unit_future_rows = defaultdict(list)

    today = datetime.today()
    current_month_start = datetime(today.year, today.month, 1)

    for r in rows:
        milestone = (r.get('Milestone') or '').strip()
        row_booking_date = parse_cell_date(r.get('SFDC Booking date'))
        row_allotment_date = parse_cell_date(r.get('Allotment Date'))
        hybrid = resolve_hybrid(milestone, row_booking_date, row_allotment_date, SLAB_SCHEDULE_DATES)
        resolved_date = None
        if hybrid:
            resolved_date, mtype, sname = hybrid
        else:
            mtype = classify(milestone)
            sname = short_name(milestone)

        demand = num(r.get('Demand Amount W/O Tax'))
        installment = num(r.get('Installment Amount'))
        bank = num(r.get('Received Amt (in Bank)'))
        cgst = num(r.get('CGST'))
        sgst = num(r.get('SGST'))
        recv_wot = bank - cgst - sgst
        outstanding = num(r.get('Outstanding 1'))
        tower_raw = (r.get('Tower') or 'Unknown').strip() or 'Unknown'
        month = to_month(r.get('Bill creation date') or r.get('SAP Booking date'))
        due_month = to_month(r.get('Due Installment Date')) or (f"{resolved_date.year}-{resolved_date.month:02d}" if resolved_date else '')
        unit = (r.get('Unit Number') or '').strip()

        for t in ('all', mtype):
            buckets[t]['totalInstallment'] += demand
            buckets[t]['totalReceivedBank'] += bank
            buckets[t]['totalReceivedWoT'] += recv_wot
            buckets[t]['totalOutstanding'] += outstanding
            buckets[t]['totalDemandWoTax'] += demand

        if demand == 0 and bank > 0:
            adv_raw_all += bank
            adv_raw[mtype] += bank
            unit_advance_pool[unit] += bank

        tk = tower_acc[tower_raw]
        tk[f'{mtype}_dem'] += demand
        tk[f'{mtype}_rec'] += recv_wot
        tk[f'{mtype}_out'] += outstanding

        if month:
            mk = monthly_acc[month]
            mk[f'{mtype}_dem'] += demand
            mk[f'{mtype}_rec'] += recv_wot
            mk['outstanding'] += outstanding

        if demand > 0:
            pass
        else:
            row_date = estimate_row_date(milestone, resolved_date, row_booking_date, row_allotment_date, sname, SLAB_SCHEDULE_DATES)
            unit_future_rows[unit].append([row_date, installment, mtype, sname, tower_raw])

    for unit, future_rows in unit_future_rows.items():
        future_rows.sort(key=lambda x: (x[0] is None, x[0] if x[0] else datetime.max))
        remaining_advance = unit_advance_pool.get(unit, 0.0)
        for row in future_rows:
            row_date, installment, mtype, sname, tower_raw = row
            net_installment = installment
            if remaining_advance > 0:
                applied = min(remaining_advance, net_installment)
                net_installment -= applied
                remaining_advance -= applied
            clamped_date = row_date
            if clamped_date is not None and clamped_date < current_month_start:
                clamped_date = current_month_start
            due_month_net = f"{clamped_date.year}-{clamped_date.month:02d}" if clamped_date else ''

            acc = milestone_acc[sname]
            acc['totalCr'] += net_installment
            acc[tower_raw] = acc.get(tower_raw, 0) + net_installment
            milestone_type[sname] = mtype
            if due_month_net:
                milestone_dates[sname].append(due_month_net)

    if buckets['all']['totalOutstanding'] > 0:
        overdue_sname = 'Overdue / Outstanding Payments'
        current_month_key = f"{current_month_start.year}-{current_month_start.month:02d}"
        milestone_acc[overdue_sname]['totalCr'] += buckets['all']['totalOutstanding']
        milestone_type[overdue_sname] = 'overdue'
        milestone_dates[overdue_sname].append(current_month_key)

    def round_bucket(b):
        return {k: round(v / 1e7, 2) for k, v in b.items()}

    kpi = {t: round_bucket(buckets[t]) for t in buckets}

    GST_RATE = 0.05
    def adv_split(raw):
        net = raw / (1 + GST_RATE)
        gst = raw - net
        return round(raw / 1e7, 2), round(net / 1e7, 2), round(gst / 1e7, 2)

    adv_all_raw, adv_all_net, adv_all_gst = adv_split(adv_raw_all)
    advance_all = {'rawCr': adv_all_raw, 'netCr': adv_all_net, 'gstCr': adv_all_gst}
    advance = {}
    for t in ('tlp', 'clp'):
        rc, nc, gc = adv_split(adv_raw[t])
        advance[t] = {'rawCr': rc, 'netCr': nc, 'gstCr': gc}

    advance_note = (f"\u26a0\ufe0f Advance Money: \u20b9{adv_all_raw} Cr received before SAP demand was raised. "
                    f"GST @5% back-calculated (\u20b9{adv_all_gst} Cr) to show net W/O Tax of \u20b9{adv_all_net} Cr.")

    # ⚠️ LOCKED: raw source data tags a handful of units with the combined label
    # 'Tower-1 & 2' (not split individually) - tower_acc is keyed by that raw
    # value, but displayed as 'Tower-1 & 2 (PH)' to flag it as a distinct small
    # bucket, not a data-entry mistake. Matches the rename in rebuild_overview.py.
    tower_order = ['Tower-1', 'Tower-2', 'Tower-1 & 2']
    TOWER_DISPLAY = {'Tower-1 & 2': 'Tower-1 & 2 (PH)'}
    towerKpi = []
    for t in tower_order:
        v = tower_acc.get(t, {})
        towerKpi.append({
            'tower': TOWER_DISPLAY.get(t, t),
            'tlp_dem': round(v.get('tlp_dem', 0) / 1e7, 2),
            'clp_dem': round(v.get('clp_dem', 0) / 1e7, 2),
            'tlp_rec': round(v.get('tlp_rec', 0) / 1e7, 2),
            'clp_rec': round(v.get('clp_rec', 0) / 1e7, 2),
            'tlp_out': round(v.get('tlp_out', 0) / 1e7, 2),
            'clp_out': round(v.get('clp_out', 0) / 1e7, 2),
        })
    for t, v in tower_acc.items():
        if t not in tower_order:
            towerKpi.append({
                'tower': t,
                'tlp_dem': round(v.get('tlp_dem', 0) / 1e7, 2), 'clp_dem': round(v.get('clp_dem', 0) / 1e7, 2),
                'tlp_rec': round(v.get('tlp_rec', 0) / 1e7, 2), 'clp_rec': round(v.get('clp_rec', 0) / 1e7, 2),
                'tlp_out': round(v.get('tlp_out', 0) / 1e7, 2), 'clp_out': round(v.get('clp_out', 0) / 1e7, 2),
            })

    monthlyTrend = []
    for m in sorted(monthly_acc.keys()):
        v = monthly_acc[m]
        dem = v.get('tlp_dem', 0) + v.get('clp_dem', 0)
        rec = v.get('tlp_rec', 0) + v.get('clp_rec', 0)
        monthlyTrend.append({
            'month': m,
            'label': mlabel(m),
            'tlp_dem': round(v.get('tlp_dem', 0) / 1e7, 2),
            'clp_dem': round(v.get('clp_dem', 0) / 1e7, 2),
            'tlp_rec': round(v.get('tlp_rec', 0) / 1e7, 2),
            'clp_rec': round(v.get('clp_rec', 0) / 1e7, 2),
            'dem': round(dem / 1e7, 2),
            'rec': round(rec / 1e7, 2),
            'outstanding': round(v.get('outstanding', 0) / 1e7, 2),
        })

    milestonesUpcoming = []
    for sname, acc in sorted(milestone_acc.items(), key=lambda x: -x[1]['totalCr']):
        mtype = milestone_type[sname]
        entry = {
            'name': sname,
            'type': mtype,
            'totalCr': round(acc['totalCr'] / 1e7, 2),
        }
        for raw_t, mapped_t in TOWER_MAP.items():
            entry[mapped_t] = round(acc.get(raw_t, 0) / 1e7, 2)
        for t in ['T3', 'T4', 'T5', 'T6']:
            entry[t] = 0.0
        dates = milestone_dates.get(sname, [])
        entry['expectedDate'] = max(dates) if dates else SLAB_SCHEDULE_DATES.get(sname, '')
        if mtype == 'overdue':
            entry['shortName'] = sname
        elif sname.startswith('Before ') or sname.startswith('After '):
            entry['shortName'] = sname
        else:
            prefix = 'CLP' if mtype == 'clp' else 'TLP'
            entry['shortName'] = f"{prefix} \u2014 {sname}"
        milestonesUpcoming.append(entry)

    out = {
        'kpi': kpi,
        'gstRate': GST_RATE,
        'advanceNote': advance_note,
        'advance': advance,
        'advance_all': advance_all,
        'milestonesUpcoming': milestonesUpcoming,
        'monthlyTrend': monthlyTrend,
        'towerKpi': towerKpi,
    }

    existing_path = os.path.join(BASE, OUT)
    if os.path.exists(existing_path):
        existing = json.load(open(existing_path))
        if existing.get('projectMeta'):
            out['projectMeta'] = existing['projectMeta']

    json.dump(out, open(existing_path, 'w'), separators=(',', ':'))

    print("\n\u2705 Done! trump_dapp_kpi.json rebuilt.")
    print(f"  kpi.all: {kpi['all']}")
    print(f"  kpi.tlp: {kpi['tlp']}")
    print(f"  kpi.clp: {kpi['clp']}")
    print(f"  advance_all: {advance_all}")
    print(f"  milestone rows: {len(milestonesUpcoming)}")
    print(f"  monthly trend points: {len(monthlyTrend)}")


if __name__ == '__main__':
    main()
