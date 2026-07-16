#!/usr/bin/env python3
"""Rebuild the Overview sections of dashboard_data.json from PDRN+INVR (+DAPP arrays)
for Edition, Sky Arc and Trump. Preserves projectMeta, monthlyTargets, filterOptions,
workflow, P&L (untouched), and the Le Courtyard / Suites entries in areaSummary."""
import openpyxl, json, os
from datetime import datetime, date
from collections import defaultdict, Counter

BASE = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')

# project config: key in JSON, kpi section key, monthly-rate key, area source, byProject label, files
PROJECTS = [
    dict(key='SMARTWORLD THE EDITION', kpi='kpiExtra',       month='monthlyActualRates',
         area='pdrn', company='ETSY REALCO',      bpLabel='SMARTWORLD THE EDITION',
         cpLabel='Edition',  files=('edition_pdrn.XLSX','edition_invr.XLSX','edition_dapp.XLSX')),
    dict(key='SMARTWORLD SKY ARC',     kpi='skyarcKpiExtra', month='skyarcMonthlyRates',
         area='invr', company='RIVERDAY INFRA',   bpLabel='Smartworld Sky Arc',
         cpLabel='Sky Arc',  files=('skyarc_pdrn.XLSX','skyarc_invr.XLSX','skyarc_dapp.XLSX')),
    dict(key='TRUMP RESIDENCES GURGAON', kpi='trumpKpiExtra', month='trumpMonthlyRates',
         area='invr', company='TRUMP RESIDENCES',  bpLabel='Trump Residences Gurgaon',
         cpLabel='Trump',    files=('trump_pdrn.XLSX','trump_invr.XLSX','trump_dapp.XLSX')),
]

def load(fn):
    wb=openpyxl.load_workbook(os.path.join(BASE,fn),read_only=True,data_only=True); ws=wb.active
    h=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
    return [{h[i]:r[i].value for i in range(len(h))} for r in ws.iter_rows(min_row=2)]

def num(v):
    try: return float(v) if v not in (None,'') else 0.0
    except: return 0.0
def s(v): return str(v).strip() if v not in (None,'') else ''
def dt(v):
    if v is None: return None
    if isinstance(v,(datetime,date)): return v.strftime('%Y-%m-%d')
    return str(v)[:10]
def fy(y,m): return f'FY{y}-{str(y+1)[2:]}' if m>=4 else f'FY{y-1}-{str(y)[2:]}'
def mlabel(m):
    try: return datetime.strptime(m,'%Y-%m').strftime("%b'%y")
    except: return m

def parse_pdrn(rows, proj, company):
    out=[]
    for r in rows:
        d=dt(r.get('SFDC Booking Date')); month='';yr=0;mn=0
        if d:
            try: p=datetime.strptime(d[:10],'%Y-%m-%d'); month=f"{p.year}-{p.month:02d}"; yr=p.year; mn=p.month
            except: pass
        bs=[v for k,v in r.items() if k=='Booking Status']
        raw=s(bs[-1]) if bs else s(r.get('Booking Status'))
        # Match existing data convention: cancelled -> CANCELLED, blank -> excluded,
        # everything else (incl. TEMPORARY SURRENDER) counts as booked/ACTIVE.
        st='CANCELLED' if 'CANCEL' in raw.upper() else ('' if raw=='' else 'ACTIVE')
        out.append(dict(company=s(r.get('Company Name')),companyNorm=company,project=proj,
            bhkFull=s(r.get('BHK')),bhk=s(r.get('BHK')),status=st,bookingMonth=month,bookingYear=yr,
            bookingDate=d,bookingFY=fy(yr,mn) if yr else '',unit=s(r.get('Unit No.')),
            broker=s(r.get('Broker Code')),brokerName=s(r.get('Broker Name (SFDC)')),
            bsp=num(r.get('Total BSP Net Value')),tcv=num(r.get('TCV (With Tax)')),
            demand=num(r.get('Total Demand Amount')),received=num(r.get('Total Received')),
            superArea=num(r.get('Super Area')),carpet=num(r.get('Carpet')),
            customer=s(r.get('Latest Customer Name')),paymentPlan=s(r.get('Payment Plan Name')),
            loanStatus=s(r.get('Loan Status')),tower=s(r.get('Tower')),floor=s(r.get('Floor')),
            cancelDate=dt(r.get('Cancellation Date')),cancelReason=s(r.get('Cancellation Reason')) or 'Not specified'))
    return out

def parse_invr(rows, proj, company):
    out=[]
    for r in rows:
        unit=s(r.get('Unit Description')) or s(r.get('            Unit Number') or r.get('Unit Number') or '')
        out.append(dict(project=proj,companyNorm=company,unit=unit,bhk=s(r.get('BHK')),
            status=s(r.get('Status')),tower=s(r.get('Tower')),floor=s(r.get('Floor')),
            superArea=num(r.get('Total Super Area')) or num(r.get('Super Builtup Area')),
            carpetArea=num(r.get('Carpet Area')),bsp=num(r.get('Basic Price'))))
    return out

def parse_dapp(rows, proj, company):
    out=[]
    for r in rows:
        d=dt(r.get('Bill creation date') or r.get('SAP Booking date')); month=''
        if d:
            try: p=datetime.strptime(d[:10],'%Y-%m-%d'); month=f"{p.year}-{p.month:02d}"
            except: pass
        out.append(dict(project=s(r.get('Project Name')) or proj,companyNorm=company,
            unit=s(r.get('Unit Number')),customer=s(r.get('Sold to Party Name') or r.get('Customer Name (Payer)')),
            billMonth=month,demand=num(r.get('Total Demand after adj of Credit') or r.get('Total Due Amount With Tax')),
            received=num(r.get('Received Amt (in Bank)')),outstanding=num(r.get('Outstanding 1') or r.get('Outstanding Amount'))))
    return out

def main():
    jp=os.path.join(BASE,'dashboard_data.json')
    d=json.load(open(jp))
    old={p['kpi']:dict(d.get(p['kpi'],{})) for p in PROJECTS}

    all_pdrn_other=[r for r in d['pdrn'] if r.get('project') not in [p['key'] for p in PROJECTS]]
    all_invr_other=[r for r in d['invr'] if r.get('project') not in [p['key'] for p in PROJECTS]]
    all_dapp_other=[r for r in d['dapp'] if r.get('project') not in [p['key'] for p in PROJECTS]]
    new_pdrn=list(all_pdrn_other); new_invr=list(all_invr_other); new_dapp=list(all_dapp_other)

    parsed={}
    for P in PROJECTS:
        pf,inf,df=P['files']
        pd=parse_pdrn(load(pf),P['key'],P['company'])
        iv=parse_invr(load(inf),P['key'],P['company'])
        da=parse_dapp(load(df),P['key'],P['company'])
        parsed[P['key']]=(pd,iv,da)
        new_pdrn+=pd; new_invr+=iv; new_dapp+=da

        act=[r for r in pd if r['status']=='ACTIVE']; can=[r for r in pd if r['status']=='CANCELLED']
        invb=[r for r in iv if r['status']=='Booked']; inva=[r for r in iv if r['status']=='Available']
        bsp=sum(r['bsp'] for r in act); tcv=sum(r['tcv'] for r in act); cbsp=sum(r['bsp'] for r in can)
        if P['area']=='pdrn':
            barea=sum(r['superArea'] for r in act); carea=sum(r['carpet'] for r in act); bunits=len(act)
        else:
            barea=sum(r['superArea'] for r in invb); carea=sum(r['carpetArea'] for r in invb); bunits=len(invb)
        avail_area=sum(r['superArea'] for r in inva); avail_units=len(inva)
        rate=round(bsp/barea) if barea else 0

        # kpiExtra - merge (not replace) so extended fields some projects carry
        # (e.g. Sky Arc's totalProjCr/soldBSPCr/unsoldValueCr/collectedCr/etc,
        # computed separately via the INVR-based Approach 2 formula) are preserved.
        d[P['kpi']] = {**d.get(P['kpi'], {}), **dict(totalBSPCr=round(bsp/1e7,1),totalTCVCr=round(tcv/1e7,1),
            bookedAreaSqft=round(barea),carpetAreaSqft=round(carea),
            cancelledBSPCr=round(cbsp/1e7,1),avgRatePerSqft=rate)}

        # monthly rate series (active bsp/area by booking month)
        mm=defaultdict(lambda:[0.0,0.0])
        for r in act:
            if r['bookingMonth']: mm[r['bookingMonth']][0]+=r['bsp']; mm[r['bookingMonth']][1]+=r['superArea']
        d[P['month']]={mlabel(k):(round(v[0]/v[1]) if v[1] else 0) for k,v in sorted(mm.items())}

        # areaSummary byProject entry
        d['areaSummary']['byProject']=[x for x in d['areaSummary']['byProject'] if x.get('project')!=P['bpLabel']]
        d['areaSummary']['byProject'].append(dict(project=P['bpLabel'],bookedArea=round(barea),
            availableArea=round(avail_area),bookedUnits=bunits,availUnits=avail_units,avgPricePerSqft=rate))

        # towerData
        d['towerData']=[t for t in d['towerData'] if t.get('project')!=P['key']]
        tw=defaultdict(lambda:dict(booked=0,cancelled=0,avail=0,ba=0.0,ca=0.0,cp=0.0,bsp=0.0))
        for r in inva:
            if r['tower']: tw[r['tower']]['avail']+=1
        for r in pd:
            t=r['tower'] or 'Unknown'
            if r['status']=='ACTIVE': tw[t]['booked']+=1; tw[t]['ba']+=r['superArea']; tw[t]['cp']+=r['carpet']; tw[t]['bsp']+=r['bsp']
            elif r['status']=='CANCELLED': tw[t]['cancelled']+=1; tw[t]['ca']+=r['superArea']
        for t,v in sorted(tw.items()):
            tot=v['booked']+v['cancelled']+v['avail']
            d['towerData'].append(dict(tower=t,project=P['key'],booked=v['booked'],cancelled=v['cancelled'],
                bookedArea=round(v['ba']),cancelledArea=round(v['ca']),carpetArea=round(v['cp']),
                pricePerSqft=round(v['bsp']/v['ba']) if v['ba']>0 else 0,totalBSPCr=round(v['bsp']/1e7,1),
                total=tot,available=v['avail'],pctSold=round(v['booked']/tot*100) if tot>0 else 0))

        # cpVsDirect
        cpu=cpb=du=db=0
        for r in act:
            if r['brokerName']: cpu+=1; cpb+=r['bsp']
            else: du+=1; db+=r['bsp']
        d['cpVsDirect']=[x for x in d.get('cpVsDirect',[]) if x.get('label')!=P['cpLabel']]
        d['cpVsDirect'].append(dict(name=P['bpLabel'],label=P['cpLabel'],cp=cpu,direct=du,
            cpBSPCr=round(cpb/1e7,1),directBSPCr=round(db/1e7,1)))

        # projBrokers / projTypologies / brokerMap (per project)
        bc=Counter(r['brokerName'] for r in pd if r['brokerName'])
        d['projBrokers'][P['key']]=[b for b,_ in bc.most_common()]
        seen=[];st_=set()
        for r in pd:
            b=r['bhkFull'] or r['bhk']
            if b and b not in st_: st_.add(b); seen.append(b)
        d['projTypologies'][P['key']]=seen

    d['pdrn']=new_pdrn; d['invr']=new_invr; d['dapp']=new_dapp

    # brokerMap across the 3 projects
    bm={}
    for P in PROJECTS:
        for r in parsed[P['key']][0]:
            if r['broker'] and r['brokerName']: bm[r['broker']]=r['brokerName']
    d['brokerMap']=bm

    # salesVsRefund: all-projects monthly (active bsp / cancelled bsp / counts)
    mo=defaultdict(lambda:dict(bsp=0.0,cbsp=0.0,bc=0,cc=0))
    for r in new_pdrn:
        m=r.get('bookingMonth')
        if not m: continue
        if r['status']=='ACTIVE': mo[m]['bsp']+=r['bsp']; mo[m]['bc']+=1
        elif r['status']=='CANCELLED': mo[m]['cbsp']+=r['bsp']; mo[m]['cc']+=1
    d['salesVsRefund']=[dict(month=mlabel(m),bspCr=round(v['bsp']/1e7,1),cancelledBSPCr=round(v['cbsp']/1e7,1),
        refundCr=0.0,bookCount=v['bc'],cancelCount=v['cc']) for m,v in sorted(mo.items())]

    # cancelledUnitStatus: consistent method across 3 (rebooked = cancelled unit now Booked in INVR)
    today=datetime.today(); byproj=[]; allvac=[]; allreb=[]; tot_c=0
    for P in PROJECTS:
        pd,iv,_=parsed[P['key']]
        booked={r['unit'] for r in iv if r['status']=='Booked'}
        can=[r for r in pd if r['status']=='CANCELLED']; tot_c+=len(can)
        vac=[];reb=[]
        for r in can:
            days=0
            if r['cancelDate']:
                try: days=(today-datetime.strptime(str(r['cancelDate'])[:10],'%Y-%m-%d')).days
                except: pass
            o=dict(unit=r['unit'],project=P['key'],projectLabel=P['cpLabel'],tower=r['tower'],
                bhk=r['bhkFull'] or r['bhk'],daysVacant=days,cancelDate=str(r['cancelDate'])[:10] if r['cancelDate'] else '',
                bspCr=round(r['bsp']/1e7,2),cancelReason=r['cancelReason'])
            (reb if r['unit'] in booked else vac).append(o)
        vac.sort(key=lambda x:-x['daysVacant'])
        byproj.append(dict(project=P['cpLabel'],rebooked=len(reb),vacant=len(vac),
            avgVacantDays=round(sum(u['daysVacant'] for u in vac)/len(vac)) if vac else 0))
        allvac+=vac; allreb+=reb
    reb_c=len(allreb); vac_c=len(allvac); bk={'0–30 days':0,'31–90 days':0,'91–180 days':0,'180+ days':0}
    for u in allvac:
        dv=u['daysVacant']
        bk['0–30 days' if dv<=30 else '31–90 days' if dv<=90 else '91–180 days' if dv<=180 else '180+ days']+=1
    d['cancelledUnitStatus']=dict(
        summary=dict(totalCancelled=tot_c,rebooked=reb_c,stillVacant=vac_c,rebookedPct=round(reb_c/tot_c*100) if tot_c else 0),
        buckets=[{'label':'0–30 days','count':bk['0–30 days'],'color':'#00bcd4'},
                 {'label':'31–90 days','count':bk['31–90 days'],'color':'#f59e0b'},
                 {'label':'91–180 days','count':bk['91–180 days'],'color':'#ef4444'},
                 {'label':'180+ days','count':bk['180+ days'],'color':'#7c3aed'}],
        byProject=byproj,vacantUnits=allvac,rebookedUnits=allreb)

    # top-level areaSummary aggregates (all byProject)
    bp=d['areaSummary']['byProject']
    tba=sum(x['bookedArea'] for x in bp); taa=sum(x['availableArea'] for x in bp)
    tbsp=0.0
    for P in PROJECTS: tbsp+=sum(r['bsp'] for r in parsed[P['key']][0] if r['status']=='ACTIVE')
    d['areaSummary']['bookedArea']=round(tba); d['areaSummary']['availableArea']=round(taa)
    d['areaSummary']['avgPricePerSqft']=round(tbsp/ (sum(sum(r['superArea'] for r in parsed[P['key']][0] if r['status']=='ACTIVE') for P in PROJECTS) or 1))

    json.dump(d,open(jp,'w'),separators=(',',':'))

    # before/after report
    print("\n==== kpiExtra BEFORE -> AFTER ====")
    for P in PROJECTS:
        o=old[P['kpi']]; n=d[P['kpi']]
        print(f"\n{P['cpLabel']} ({P['kpi']}):")
        for k in n:
            mark='' if o.get(k)==n[k] else '  <-- changed'
            print(f"   {k:16} {str(o.get(k)):>12} -> {str(n[k]):>12}{mark}")

if __name__=='__main__': main()
