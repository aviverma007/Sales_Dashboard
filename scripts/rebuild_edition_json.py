#!/usr/bin/env python3
"""
rebuild_edition_json.py
-----------------------
Reads edition_pdrn.XLSX, edition_invr.XLSX, edition_dapp.XLSX from public/data/
and rebuilds the Edition-specific sections of dashboard_data.json.
All other project data (Sky Arc, Trump, Le Courtyard) is preserved exactly.

Usage:
    python3 scripts/rebuild_edition_json.py
"""

import openpyxl
import json
import os
from datetime import datetime, date
from collections import defaultdict

BASE = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
EDITION = 'SMARTWORLD THE EDITION'
COMPANY_NORM = 'ETSY REALCO'


def load_sheet(filename):
    wb = openpyxl.load_workbook(os.path.join(BASE, filename))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2):
        rows.append({headers[i]: row[i].value for i in range(len(headers))})
    return rows


def safe_num(v):
    try:
        return float(v) if v not in (None, '') else 0.0
    except Exception:
        return 0.0


def safe_str(v):
    return str(v).strip() if v not in (None, '') else ''


def to_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    try:
        return str(v)[:10]
    except Exception:
        return None


def fy_label(year, month_num):
    if month_num >= 4:
        return f'FY{year}-{str(year+1)[2:]}'
    else:
        return f'FY{year-1}-{str(year)[2:]}'


# ─────────────────────────────────────────────
# Parse PDRN
# ─────────────────────────────────────────────
def parse_pdrn(rows):
    records = []
    for r in rows:
        bd = r.get('SFDC Booking Date')
        dt = to_date(bd)
        month = ''
        year = 0
        month_num = 0
        if dt:
            try:
                d = datetime.strptime(dt[:10], '%Y-%m-%d')
                month = f"{d.year}-{str(d.month).padStart(2,'0') if hasattr(str(d.month),'padStart') else str(d.month).zfill(2)}"
                year = d.year
                month_num = d.month
            except Exception:
                pass

        # Determine status from last 'Booking Status' column (there are two)
        booking_status_vals = [v for k, v in r.items() if k == 'Booking Status']
        status_raw = booking_status_vals[-1] if booking_status_vals else (booking_status_vals[0] if booking_status_vals else '')
        if not status_raw:
            status_raw = r.get('Booking Status', '')
        status = 'ACTIVE' if safe_str(status_raw) == 'ACTIVE' else ('CANCELLED' if 'CANCEL' in safe_str(status_raw).upper() else safe_str(status_raw))

        records.append({
            'company': safe_str(r.get('Company Name')),
            'companyNorm': COMPANY_NORM,
            'project': EDITION,
            'bhkFull': safe_str(r.get('BHK')),
            'bhk': safe_str(r.get('BHK')),
            'status': status,
            'bookingMonth': month,
            'bookingYear': year,
            'bookingDate': dt,
            'bookingFY': fy_label(year, month_num) if year else '',
            'unit': safe_str(r.get('Unit No.')),
            'broker': safe_str(r.get('Broker Code')),
            'brokerName': safe_str(r.get('Broker Name (SFDC)')),
            'bsp': safe_num(r.get('Total BSP Net Value')),
            'tcv': safe_num(r.get('TCV (With Tax)')),
            'demand': safe_num(r.get('Total Demand Amount')),
            'received': safe_num(r.get('Total Received')),
            'superArea': safe_num(r.get('Super Area')),
            'carpet': safe_num(r.get('Carpet')),
            'customer': safe_str(r.get('Latest Customer Name')),
            'paymentPlan': safe_str(r.get('Payment Plan Name')),
            'loanStatus': safe_str(r.get('Loan Status')),
            'tower': safe_str(r.get('Tower')),
            'floor': safe_str(r.get('Floor')),
            'cancelDate': to_date(r.get('Cancellation Date')),
            'cancelReason': safe_str(r.get('Cancellation Reason')) or 'Not specified',
        })
    return records


# ─────────────────────────────────────────────
# Parse INVR
# ─────────────────────────────────────────────
def parse_invr(rows):
    records = []
    for r in rows:
        unit_raw = r.get('            Unit Number') or r.get('Unit Number') or ''
        unit = safe_str(unit_raw)
        # Extract short unit description (T6-904 style)
        unit_desc = safe_str(r.get('Unit Description'))
        records.append({
            'project': EDITION,
            'companyNorm': COMPANY_NORM,
            'unit': unit_desc or unit,
            'bhk': safe_str(r.get('BHK')),
            'status': safe_str(r.get('Status')),
            'tower': safe_str(r.get('Tower')),
            'floor': safe_str(r.get('Floor')),
            'superArea': safe_num(r.get('Total Super Area')) or safe_num(r.get('Super Builtup Area')),
            'carpetArea': safe_num(r.get('Carpet Area')),
            'bsp': safe_num(r.get('Basic Price')),
        })
    return records


# ─────────────────────────────────────────────
# Parse DAPP
# ─────────────────────────────────────────────
def parse_dapp(rows):
    records = []
    for r in rows:
        bd = r.get('Bill creation date') or r.get('SAP Booking date')
        dt = to_date(bd)
        month = ''
        if dt:
            try:
                d = datetime.strptime(dt[:10], '%Y-%m-%d')
                month = f"{d.year}-{str(d.month).zfill(2)}"
            except Exception:
                pass

        proj = safe_str(r.get('Project Name')) or EDITION
        records.append({
            'project': proj,
            'companyNorm': COMPANY_NORM,
            'unit': safe_str(r.get('Unit Number')),
            'customer': safe_str(r.get('Sold to Party Name') or r.get('Customer Name (Payer)')),
            'billMonth': month,
            'demand': safe_num(r.get('Total Due Amount With Tax') or r.get('Installment Amount')),
            'received': safe_num(r.get('Received Amount')),
            'outstanding': safe_num(r.get('Outstanding Amount')),
        })
    return records


# ─────────────────────────────────────────────
# Derived sections for Edition
# ─────────────────────────────────────────────
def build_proj_brokers(pdrn_records):
    counts = defaultdict(int)
    for r in pdrn_records:
        if r['brokerName']:
            counts[r['brokerName']] += 1
    return [b for b, _ in sorted(counts.items(), key=lambda x: -x[1])]


def build_proj_typologies(pdrn_records):
    seen = []
    s = set()
    for r in pdrn_records:
        bhk = r['bhkFull'] or r['bhk']
        if bhk and bhk not in s:
            s.add(bhk)
            seen.append(bhk)
    return seen


def build_tower_data(pdrn_records, invr_records=[]):
    # Count available from invr
    inv_avail = defaultdict(int)
    for r in invr_records:
        if r['status'] == 'Available' and r.get('tower'):
            inv_avail[r['tower']] += 1

    towers = defaultdict(lambda: {'booked': 0, 'cancelled': 0, 'avail': 0, 'bookedArea': 0.0,
                                   'cancelledArea': 0.0, 'carpetArea': 0.0, 'bspTotal': 0.0})
    for r in pdrn_records:
        t = r['tower'] or 'Unknown'
        area = r['superArea'] or 0
        carpet = r['carpet'] or 0
        bsp = r['bsp'] or 0
        if r['status'] == 'ACTIVE':
            towers[t]['booked'] += 1
            towers[t]['bookedArea'] += area
            towers[t]['carpetArea'] += carpet
            towers[t]['bspTotal'] += bsp
        elif r['status'] == 'CANCELLED':
            towers[t]['cancelled'] += 1
            towers[t]['cancelledArea'] += area
    for t, cnt in inv_avail.items():
        towers[t]['avail'] = cnt
    result = []
    for t, v in sorted(towers.items()):
        price = round(v['bspTotal'] / v['bookedArea']) if v['bookedArea'] > 0 else 0
        total = v['booked'] + v['cancelled'] + v.get('avail', 0)
        result.append({
            'tower': t,
            'project': EDITION,
            'booked': v['booked'],
            'cancelled': v['cancelled'],
            'bookedArea': round(v['bookedArea']),
            'cancelledArea': round(v['cancelledArea']),
            'carpetArea': round(v['carpetArea']),
            'pricePerSqft': price,
            'totalBSPCr': round(v['bspTotal'] / 1e7, 1),
            'total': total,
            'available': v.get('avail', 0),
            'pctSold': round(v['booked'] / total * 100) if total > 0 else 0,
        })
    return result


def build_sales_vs_refund(pdrn_records):
    monthly = defaultdict(lambda: {'bsp': 0.0, 'cancelledBSP': 0.0, 'refund': 0.0,
                                    'bookCount': 0, 'cancelCount': 0})
    for r in pdrn_records:
        m = r['bookingMonth']
        if not m:
            continue
        if r['status'] == 'ACTIVE':
            monthly[m]['bsp'] += r['bsp']
            monthly[m]['bookCount'] += 1
        elif r['status'] == 'CANCELLED':
            monthly[m]['cancelledBSP'] += r['bsp']
            monthly[m]['cancelCount'] += 1

    result = []
    for m in sorted(monthly.keys()):
        v = monthly[m]
        try:
            d = datetime.strptime(m, '%Y-%m')
            label = d.strftime("%b'%y")
        except Exception:
            label = m
        result.append({
            'month': label,
            'bspCr': round(v['bsp'] / 1e7, 1),
            'cancelledBSPCr': round(v['cancelledBSP'] / 1e7, 1),
            'refundCr': 0.0,
            'bookCount': v['bookCount'],
            'cancelCount': v['cancelCount'],
        })
    return result


def build_cp_vs_direct(pdrn_records):
    cp_units, direct_units = 0, 0
    cp_bsp, direct_bsp = 0.0, 0.0
    for r in pdrn_records:
        if r['status'] != 'ACTIVE':
            continue
        if r['brokerName']:
            cp_units += 1
            cp_bsp += r['bsp']
        else:
            direct_units += 1
            direct_bsp += r['bsp']
    return {
        'name': EDITION,
        'label': 'Edition',
        'cp': cp_units,
        'direct': direct_units,
        'cpBSPCr': round(cp_bsp / 1e7, 1),
        'directBSPCr': round(direct_bsp / 1e7, 1),
    }


def build_area_summary(pdrn_records, invr_records):
    active = [r for r in pdrn_records if r['status'] == 'ACTIVE']
    avail = [r for r in invr_records if r['status'] == 'Available']
    booked_area = sum(r['superArea'] for r in active)
    avail_area = sum(r['superArea'] for r in avail)
    total_bsp = sum(r['bsp'] for r in active)
    avg_price = round(total_bsp / booked_area) if booked_area > 0 else 0
    return {
        'project': EDITION,
        'bookedArea': round(booked_area),
        'availableArea': round(avail_area),
        'bookedUnits': len(active),
        'availUnits': len(avail),
        'avgPricePerSqft': avg_price,
    }


def build_kpi_extra(pdrn_records):
    active = [r for r in pdrn_records if r['status'] == 'ACTIVE']
    cancelled = [r for r in pdrn_records if r['status'] == 'CANCELLED']
    total_bsp = sum(r['bsp'] for r in active)
    total_tcv = sum(r['tcv'] for r in active)
    booked_area = sum(r['superArea'] for r in active)
    carpet_area = sum(r['carpet'] for r in active)
    cancelled_bsp = sum(r['bsp'] for r in cancelled)
    avg_rate = round(total_bsp / booked_area) if booked_area > 0 else 0
    return {
        'totalBSPCr': round(total_bsp / 1e7, 1),
        'totalTCVCr': round(total_tcv / 1e7, 1),
        'bookedAreaSqft': round(booked_area),
        'carpetAreaSqft': round(carpet_area),
        'cancelledBSPCr': round(cancelled_bsp / 1e7, 1),
        'avgRatePerSqft': avg_rate,
    }


def build_cancelled_status(pdrn_records, invr_records):
    from datetime import datetime
    PROJ_LABEL = 'Edition'
    today = datetime.today()

    # Units currently booked in invr = rebooked after cancellation
    invr_booked = {r['unit'] for r in invr_records if r['status'] == 'Booked'}

    cancelled = [r for r in pdrn_records if r['status'] == 'CANCELLED']
    vacant_units = []
    rebooked_units = []

    for r in cancelled:
        unit = r['unit']
        cancel_date = r.get('cancelDate') or ''
        bsp_cr = round((r['bsp'] or 0) / 1e7, 2)
        days_vacant = 0
        if cancel_date:
            try:
                d = datetime.strptime(str(cancel_date)[:10], '%Y-%m-%d')
                days_vacant = (today - d).days
            except Exception:
                pass

        unit_obj = {
            'unit': unit,
            'project': EDITION,
            'projectLabel': PROJ_LABEL,
            'tower': r['tower'],
            'bhk': r['bhkFull'] or r['bhk'],
            'daysVacant': days_vacant,
            'cancelDate': str(cancel_date)[:10] if cancel_date else '',
            'bspCr': bsp_cr,
            'cancelReason': r.get('cancelReason') or 'Not specified',
        }

        if unit in invr_booked:
            rebooked_units.append(unit_obj)
        else:
            vacant_units.append(unit_obj)

    # Sort vacant by daysVacant desc (most urgent first)
    vacant_units.sort(key=lambda x: -x['daysVacant'])
    total = len(cancelled)
    rebooked_count = len(rebooked_units)
    vacant_count = len(vacant_units)

    # Bucket by daysVacant for vacant units
    bucket_counts = {'0–30 days': 0, '31–90 days': 0, '91–180 days': 0, '180+ days': 0}
    for u in vacant_units:
        d = u['daysVacant']
        if d <= 30:   bucket_counts['0–30 days'] += 1
        elif d <= 90:  bucket_counts['31–90 days'] += 1
        elif d <= 180: bucket_counts['91–180 days'] += 1
        else:          bucket_counts['180+ days'] += 1

    by_project = [{
        'project': PROJ_LABEL,
        'rebooked': rebooked_count,
        'vacant': vacant_count,
        'avgVacantDays': round(sum(u['daysVacant'] for u in vacant_units) / vacant_count) if vacant_count else 0,
    }]

    return {
        'summary': {
            'totalCancelled': total,
            'rebooked': rebooked_count,
            'stillVacant': vacant_count,
            'rebookedPct': round(rebooked_count / total * 100) if total > 0 else 0,
        },
        'buckets': [
            {'label': '0–30 days', 'count': bucket_counts['0–30 days'], 'color': '#00bcd4'},
            {'label': '31–90 days', 'count': bucket_counts['31–90 days'], 'color': '#f59e0b'},
            {'label': '91–180 days', 'count': bucket_counts['91–180 days'], 'color': '#ef4444'},
            {'label': '180+ days', 'count': bucket_counts['180+ days'], 'color': '#7c3aed'},
        ],
        'byProject': by_project,
        'vacantUnits': vacant_units,
        'rebookedUnits': rebooked_units,
    }


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("Loading Excel files...")
    pdrn_raw = load_sheet('edition_pdrn.XLSX')
    invr_raw = load_sheet('edition_invr.XLSX')
    dapp_raw = load_sheet('edition_dapp.XLSX')

    print(f"  PDRN rows: {len(pdrn_raw)}")
    print(f"  INVR rows: {len(invr_raw)}")
    print(f"  DAPP rows: {len(dapp_raw)}")

    print("Parsing...")
    edition_pdrn = parse_pdrn(pdrn_raw)
    edition_invr = parse_invr(invr_raw)
    edition_dapp = parse_dapp(dapp_raw)

    print("Loading existing JSON...")
    json_path = os.path.join(BASE, 'dashboard_data.json')
    with open(json_path) as f:
        d = json.load(f)

    print("Rebuilding Edition sections...")

    # 1. Replace pdrn records for Edition
    other_pdrn = [r for r in d['pdrn'] if r.get('project') != EDITION]
    d['pdrn'] = other_pdrn + edition_pdrn
    print(f"  pdrn: {len(other_pdrn)} other + {len(edition_pdrn)} edition = {len(d['pdrn'])} total")

    # 2. Replace invr records for Edition
    other_invr = [r for r in d['invr'] if r.get('project') != EDITION]
    d['invr'] = other_invr + edition_invr
    print(f"  invr: {len(other_invr)} other + {len(edition_invr)} edition = {len(d['invr'])} total")

    # 3. Replace dapp records for Edition
    other_dapp = [r for r in d['dapp'] if r.get('project') != EDITION]
    d['dapp'] = other_dapp + edition_dapp
    print(f"  dapp: {len(other_dapp)} other + {len(edition_dapp)} edition = {len(d['dapp'])} total")

    # 4. projBrokers
    d['projBrokers'][EDITION] = build_proj_brokers(edition_pdrn)
    print(f"  projBrokers[Edition]: {len(d['projBrokers'][EDITION])} brokers")

    # 5. projTypologies
    d['projTypologies'][EDITION] = build_proj_typologies(edition_pdrn)
    print(f"  projTypologies[Edition]: {d['projTypologies'][EDITION]}")

    # 6. towerData - replace Edition towers
    other_towers = [r for r in d['towerData'] if r.get('project') != EDITION]
    edition_towers = build_tower_data(edition_pdrn, edition_invr)
    d['towerData'] = other_towers + edition_towers
    print(f"  towerData: {len(edition_towers)} Edition towers rebuilt")

    # 7. salesVsRefund (Edition-specific; if stored per-project replace, else skip)
    # This is a global list — rebuild Edition portion
    other_svr = [r for r in d.get('salesVsRefund', []) if r.get('project') and r['project'] != EDITION]
    if not other_svr:
        # It's global not per-project, rebuild entirely from Edition data
        d['salesVsRefund'] = build_sales_vs_refund(edition_pdrn)
    
    # 8. cpVsDirect - replace Edition entry
    edition_cpd = build_cp_vs_direct(edition_pdrn)
    d['cpVsDirect'] = [r for r in d.get('cpVsDirect', []) if r.get('name') != EDITION] + [edition_cpd]

    # 9. areaSummary - replace Edition byProject entry
    edition_area = build_area_summary(edition_pdrn, edition_invr)
    d['areaSummary']['byProject'] = [r for r in d['areaSummary'].get('byProject', []) if r.get('project') != EDITION] + [edition_area]

    # 10. kpiExtra - rebuild from Edition data (this was Edition-only)
    d['kpiExtra'] = build_kpi_extra(edition_pdrn)

    # 11. cancelledUnitStatus - rebuild from Edition data
    d['cancelledUnitStatus'] = build_cancelled_status(edition_pdrn, edition_invr)

    # 12. brokerMap - rebuild from Edition pdrn
    broker_map = {}
    for r in edition_pdrn:
        if r['broker'] and r['brokerName']:
            broker_map[r['broker']] = r['brokerName']
    d['brokerMap'] = broker_map

    print("Writing JSON...")
    with open(json_path, 'w') as f:
        json.dump(d, f, separators=(',', ':'))

    print(f"\n✅ Done! dashboard_data.json updated.")
    print(f"   Edition active bookings : {sum(1 for r in edition_pdrn if r['status']=='ACTIVE')}")
    print(f"   Edition cancelled       : {sum(1 for r in edition_pdrn if r['status']=='CANCELLED')}")
    print(f"   Edition invr units      : {len(edition_invr)}")
    print(f"   Edition dapp records    : {len(edition_dapp)}")
    print(f"   Edition brokers         : {len(d['projBrokers'][EDITION])}")
    print(f"   Total pdrn records      : {len(d['pdrn'])}")


if __name__ == '__main__':
    main()
